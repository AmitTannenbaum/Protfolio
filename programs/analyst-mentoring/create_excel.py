import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Style definitions ---
title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Calibri", size=11, bold=True)
normal_font = Font(name="Calibri", size=11)
small_font = Font(name="Calibri", size=10, italic=True, color="666666")

title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
light_blue = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
light_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
light_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

rtl_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
rtl_title = Alignment(horizontal="right", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

def style_row(ws, row, font, fill, alignment, border=thin_border):
    for cell in ws[row]:
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border

def write_row(ws, row, values, font=normal_font, fill=white_fill, alignment=rtl_align):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = thin_border

# ============================================================
# Sheet 1: תוכנית מפגשים
# ============================================================
ws1 = wb.active
ws1.title = "תוכנית מפגשים"
ws1.sheet_view.rightToLeft = True

headers = ["שלב", "שבוע", "מפגש", "נושא", "מטרה", "מה עושים", "תוצר", "סטטוס"]
col_widths = [12, 10, 10, 22, 30, 50, 40, 12]

for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title row
ws1.merge_cells("A1:H1")
ws1["A1"].value = "תוכנית ליווי אנליסט — מפגשים"
ws1["A1"].font = title_font
ws1["A1"].fill = title_fill
ws1["A1"].alignment = rtl_title

# Headers
write_row(ws1, 2, headers, font=header_font, fill=header_fill, alignment=center_align)

meetings = [
    ["שלב 1: פוזיציה וסיפור", "1", "1", "מיקוד וכיוון",
     "להגדיר בדיוק לאיזה תפקידים מכוונים",
     "מיפוי תפקידים (Product/BI/Data Analyst)\nבחירת סוג חברה\nמיפוי חוזקות מול פערים\nהגדרת אנליסט אידיאלי",
     "משפט מיצוב ברור\nרשימת פערים + תוכנית סגירה", ""],

    ["שלב 1: פוזיציה וסיפור", "2", "2", "CV + LinkedIn + נוכחות דיגיטלית",
     "להפוך את המועמד/ת ל-hireable",
     "שכתוב CV מאפס\nהתאמה ל-ATS\nבניית LinkedIn מקצועי\nסידור GitHub/Portfolio",
     "CV ברמה גבוהה\nLinkedIn מלוטש\nGitHub מסודר", ""],

    ["שלב 2: פרויקטים", "3", "3", "בחירת פרויקט + הגדרת בעיה",
     "לבחור פרויקט אמיתי עם שאלה עסקית",
     "בחירת פרויקט (Funnel/Cohort/A-B/Dashboard/RFM/Churn)\nהגדרת שאלה עסקית\nבחירת דאטאסט\nהגדרת KPIs",
     "מסמך פרויקט מלא\nדאטאסט מוכן", ""],

    ["שלב 2: פרויקטים", "4", "4", "בנייה — SQL + EDA (חלק א׳)",
     "עבודה אנליטית אמיתית",
     "כתיבת SQL מורכב (JOINs, Window Functions, CTEs)\nניתוח חקירתי עם Python/Pandas\nתיעוד התהליך",
     "קובץ SQL מתועד\nNotebook עם ניתוח", ""],

    ["שלב 2: פרויקטים", "5", "5", "בנייה — תובנות (חלק ב׳)",
     "להפוך דאטה לתובנות עסקיות",
     "סימולציה: בריף מסטייקהולדר\nהפקת תובנות — לא רק גרפים אלא ״אז מה?״\nסיכום מקצועי",
     "סיכום תובנות (PDF/Notion)\nתשובה לסטייקהולדר", ""],

    ["שלב 2: פרויקטים", "6", "6", "Dashboard + Storytelling",
     "להפוך דאטה לסיפור שמנהלים מבינים",
     "בניית Dashboard (Power BI/Tableau)\nעיצוב ויזואלי נכון\nData Storytelling\nמצגת 5-7 שקפים",
     "Dashboard מלוטש\nמצגת Executive Summary\nפרויקט ב-GitHub + Tableau Public", ""],

    ["שלב 3: הכנה לראיונות", "7", "7", "שאלות מקצועיות + Case Study",
     "לשלוט בראיונות מקצועיים",
     "שאלות SQL (Window Functions, Optimization)\nשאלות מוצר (מדדים, A/B, North Star)\nCase Study: ״ה-DAU ירד 15%״\nסטטיסטיקה (p-value, sample size)\nPython (Pandas, cleaning)\nBehavioral",
     "בנק 30+ שאלות עם תשובות\nFramework ל-Case Study\nסיפור מוכן לכל פרויקט", ""],

    ["שלב 3: הכנה לראיונות", "8", "8", "סימולציות ראיון",
     "הורדת לחץ + שיפור ביצוע",
     "ראיון מלא (45-60 דק)\nפידבק מפורט: תוכן, מבנה, ניהול זמן\nחזרה על נקודות חלשות\nתרגול ״לחשוב בקול״",
     "רשימת פערים ממוקדת\nתשובות משופרות\nביטחון גבוה יותר", ""],

    ["שלב 4: יציאה לשוק", "9", "9", "אסטרטגיית חיפוש עבודה",
     "לעבוד חכם — לא רק לשלוח CV",
     "בניית רשימת 30-50 חברות יעד\nכתיבת הודעות פנייה ל-LinkedIn\nNetworking נכון\nהכנת ״למה אתם?״ לכל חברה",
     "רשימת חברות מתועדפת\n5 הודעות פנייה מוכנות\nתוכנית פעולה שבועית", ""],

    ["שלב 4: יציאה לשוק", "10+", "שבועי", "Follow-up שבועי (30 דק)",
     "ליווי עד מציאת עבודה",
     "סקירת פעילות שבועית\nניתוח דחיות\nשיפור נקודתי\nמוטיבציה ותמיכה",
     "דוח שבועי\nשיפורים ממוקדים", ""],
]

stage_fills = {
    "שלב 1: פוזיציה וסיפור": light_blue,
    "שלב 2: פרויקטים": light_green,
    "שלב 3: הכנה לראיונות": light_yellow,
    "שלב 4: יציאה לשוק": light_orange,
}

for i, row_data in enumerate(meetings):
    r = i + 3
    fill = stage_fills.get(row_data[0], white_fill)
    write_row(ws1, r, row_data, fill=fill)

# ============================================================
# Sheet 2: פרויקטים אפשריים
# ============================================================
ws2 = wb.create_sheet("פרויקטים אפשריים")
ws2.sheet_view.rightToLeft = True

headers2 = ["פרויקט", "שאלה עסקית", "מתאים לתפקיד", "כלים", "רמת קושי", "זמן משוער"]
col_widths2 = [22, 35, 22, 30, 12, 12]

for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:F1")
ws2["A1"].value = "רעיונות לפרויקטים"
ws2["A1"].font = title_font
ws2["A1"].fill = title_fill
ws2["A1"].alignment = rtl_title

write_row(ws2, 2, headers2, font=header_font, fill=header_fill, alignment=center_align)

projects = [
    ["Funnel Analysis", "איפה המשתמשים נושרים ולמה?", "Product Analyst", "SQL, Python, Tableau", "בינוני", "2 שבועות"],
    ["Cohort & Retention", "מה גורם למשתמשים לחזור?", "Product / Growth", "SQL, Python, Power BI", "בינוני-גבוה", "2 שבועות"],
    ["A/B Testing", "האם השינוי שעשינו באמת עובד?", "Product / Data Analyst", "SQL, Python, סטטיסטיקה", "גבוה", "2-3 שבועות"],
    ["Dashboard עסקי", "מה מצב הביזנס?", "BI Analyst", "SQL, Tableau/Power BI", "בינוני", "1-2 שבועות"],
    ["RFM & Segmentation", "מי הלקוחות הכי חשובים?", "Marketing / BI", "SQL, Python, Visualization", "בינוני", "2 שבועות"],
    ["Churn Prediction", "מי עומד לעזוב ולמה?", "Data Analyst", "SQL, Python, ML basics", "גבוה", "3 שבועות"],
]

for i, row_data in enumerate(projects):
    fill = light_blue if i % 2 == 0 else white_fill
    write_row(ws2, i + 3, row_data, fill=fill)

# ============================================================
# Sheet 3: KPIs ומעקב שבועי
# ============================================================
ws3 = wb.create_sheet("מעקב שבועי")
ws3.sheet_view.rightToLeft = True

headers3 = ["שבוע", "פניות שנשלחו", "תשובות שחזרו", "% תשובות", "ראיונות טלפוניים",
            "ראיונות מקצועיים", "הצעות", "איפה נתקענו", "פעולה לשבוע הבא"]
col_widths3 = [10, 15, 15, 12, 18, 18, 10, 30, 30]

for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:I1")
ws3["A1"].value = "מעקב שבועי — חיפוש עבודה"
ws3["A1"].font = title_font
ws3["A1"].fill = title_fill
ws3["A1"].alignment = rtl_title

write_row(ws3, 2, headers3, font=header_font, fill=header_fill, alignment=center_align)

# Target row
targets = ["יעד", "20-30", "4-6", "10-20%", "3-5", "1-2", "-", "", ""]
write_row(ws3, 3, targets, font=bold_font, fill=light_yellow)

# Empty weeks for tracking
for i in range(12):
    fill = light_blue if i % 2 == 0 else white_fill
    write_row(ws3, i + 4, [f"שבוע {i+1}"] + [""] * 8, fill=fill)

# ============================================================
# Sheet 4: הכנה לראיונות
# ============================================================
ws4 = wb.create_sheet("בנק שאלות ראיון")
ws4.sheet_view.rightToLeft = True

headers4 = ["קטגוריה", "שאלה", "תשובה מוכנה", "רמת קושי", "תורגל?"]
col_widths4 = [18, 45, 50, 12, 10]

for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("A1:E1")
ws4["A1"].value = "בנק שאלות ראיון"
ws4["A1"].font = title_font
ws4["A1"].fill = title_fill
ws4["A1"].alignment = rtl_title

write_row(ws4, 2, headers4, font=header_font, fill=header_fill, alignment=center_align)

questions = [
    ["SQL", "מה ההבדל בין WHERE ל-HAVING?", "", "קל", ""],
    ["SQL", "הסבר/י Window Functions ותן/י דוגמה", "", "בינוני", ""],
    ["SQL", "איך תמצא/י כפילויות בטבלה?", "", "קל", ""],
    ["SQL", "מה ההבדל בין RANK, DENSE_RANK, ROW_NUMBER?", "", "בינוני", ""],
    ["SQL", "כתוב/י Query שמחזיר Retention לפי Cohort", "", "גבוה", ""],
    ["SQL", "איך תייעל/י Query איטי?", "", "גבוה", ""],
    ["מוצר", "איך תמדוד/י הצלחה של פיצ׳ר חדש?", "", "בינוני", ""],
    ["מוצר", "מה זה North Star Metric? תן/י דוגמה", "", "בינוני", ""],
    ["מוצר", "איך תתכנן/י A/B Test?", "", "גבוה", ""],
    ["מוצר", "מה ההבדל בין Vanity Metric ל-Actionable Metric?", "", "קל", ""],
    ["Case Study", "ה-DAU ירד ב-15% — מה עושים?", "", "גבוה", ""],
    ["Case Study", "ה-Conversion Rate ירד אחרי שינוי בדף — למה?", "", "גבוה", ""],
    ["Case Study", "המנכ״ל רוצה לדעת למה ההכנסות ירדו ברבעון — איך תחקור/י?", "", "גבוה", ""],
    ["סטטיסטיקה", "מה זה p-value ומתי התוצאה significant?", "", "בינוני", ""],
    ["סטטיסטיקה", "איך קובעים sample size ל-A/B Test?", "", "גבוה", ""],
    ["Python", "איך תנקה/י דאטאסט עם ערכים חסרים?", "", "קל", ""],
    ["Python", "הסבר/י groupby ו-merge ב-Pandas", "", "בינוני", ""],
    ["Behavioral", "ספר/י על פרויקט שהובלת — מה היה האתגר?", "", "בינוני", ""],
    ["Behavioral", "תן/י דוגמה לתובנה שהובילה לשינוי בפרודקט", "", "בינוני", ""],
    ["Behavioral", "איך את/ה מתעדף/ת כשיש הרבה בקשות?", "", "קל", ""],
]

for i, row_data in enumerate(questions):
    cat_fills = {
        "SQL": light_blue,
        "מוצר": light_green,
        "Case Study": light_orange,
        "סטטיסטיקה": light_yellow,
        "Python": light_blue,
        "Behavioral": light_green,
    }
    fill = cat_fills.get(row_data[0], white_fill)
    write_row(ws4, i + 3, row_data, fill=fill)

# ============================================================
# Sheet 5: עקרונות מנחים
# ============================================================
ws5 = wb.create_sheet("עקרונות מנחים")
ws5.sheet_view.rightToLeft = True

headers5 = ["עיקרון", "למה זה חשוב", "איך מיישמים"]
col_widths5 = [25, 40, 45]

for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells("A1:C1")
ws5["A1"].value = "עקרונות מנחים לתוכנית"
ws5["A1"].font = title_font
ws5["A1"].fill = title_fill
ws5["A1"].alignment = rtl_title

write_row(ws5, 2, headers5, font=header_font, fill=header_fill, alignment=center_align)

principles = [
    ["כל מפגש מסתיים בתוצר", "לא רק דיבור — משהו שנכנס לפורטפוליו", "מגדירים תוצר צפוי בתחילת כל מפגש ובודקים בסוף"],
    ["בונים סיפור מקצועי", "מגייסים קונים סיפור, לא רשימת סקילים", "כל פרויקט = סיפור עם בעיה, תהליך, ותובנה"],
    ["עובדים כמו בעולם אמיתי", "ההבדל בין ״למדתי״ ל-״עבדתי״", "המנטור = סטייקהולדר, המנטי = אנליסט, בריפים אמיתיים"],
    ["מדידה שבועית", "אם לא מודדים — לא משפרים", "טבלת מעקב שבועית עם KPIs ברורים"],
    ["פידבק ישיר וכנה", "שיפור אמיתי דורש כנות", "פידבק על כל תוצר + סימולציות עם הערות מפורטות"],
    ["סימולציה, לא קורס", "אף אחד לא מעסיק בגלל קורס", "הכל מבוסס על תרחישים מהעולם האמיתי"],
]

for i, row_data in enumerate(principles):
    fill = light_blue if i % 2 == 0 else white_fill
    write_row(ws5, i + 3, row_data, fill=fill)

# ============================================================
# Save
# ============================================================
output_path = "/home/user/Protfolio/programs/analyst-mentoring/analyst_mentoring_program.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
